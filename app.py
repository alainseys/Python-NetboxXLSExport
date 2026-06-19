#!/usr/bin/env python3
"""
netbox_export_to_xlsx.py

Export:
  • IP Ranges
  • IP Addresses
→ Two Excel files → Email to alain@provider.com

SMTP: mailserver-01.provider.com:25 (no auth)
"""

import argparse
import sys
from pathlib import Path
from typing import List, Dict, Any
from email.message import EmailMessage
import smtplib

import requests
import urllib3
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

# --------------------------------------------------------------------------- #
# Suppress SSL warnings
# --------------------------------------------------------------------------- #
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# --------------------------------------------------------------------------- #
# Hardcoded Settings
# --------------------------------------------------------------------------- #
EMAIL_TO = ""
SMTP_SERVER = ""
SMTP_PORT = 25
FROM_EMAIL = ""

# NetBox API Endpoints
BASE_URL = "https://localhost/"
URL_IP_RANGES = f"{BASE_URL}/api/ipam/ip-ranges/"
URL_IP_ADDRESSES = f"{BASE_URL}/api/ipam/ip-addresses/"

# Output files
OUTPUT_IP_RANGES = Path("ip_ranges.xlsx")
OUTPUT_IP_ADDRESSES = Path("ip_addresses.xlsx")


# --------------------------------------------------------------------------- #
# Safe string conversion
# --------------------------------------------------------------------------- #
def safe_string(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (str, int, float)):
        return str(value)
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, dict):
        return (
            value.get("display")
            or value.get("name")
            or value.get("label")
            or value.get("value")
            or str(value)
        )
    if isinstance(value, list):
        items = [
            item.get("display") or item.get("name") or item.get("label") or str(item)
            for item in value
        ]
        return ", ".join(filter(None, items))
    return str(value)


# --------------------------------------------------------------------------- #
# Flatten IP Range
# --------------------------------------------------------------------------- #
def flatten_ip_range(obj: Dict[str, Any]) -> Dict[str, Any]:
    flat = {
        "ID": obj.get("id"),
        "Display": obj.get("display"),
        "Start Address": obj.get("start_address"),
        "End Address": obj.get("end_address"),
        "Size": obj.get("size"),
        "Family": safe_string(obj.get("family")),
        "Status": safe_string(obj.get("status")),
        "VRF": safe_string(obj.get("vrf")),
        "Tenant": safe_string(obj.get("tenant")),
        "Role": safe_string(obj.get("role")),
        "Description": obj.get("description", ""),
        "Comments": obj.get("comments", ""),
        "Mark Utilized": obj.get("mark_utilized"),
        "Created": obj.get("created"),
        "Last Updated": obj.get("last_updated"),
    }
    flat["Tags"] = safe_string(obj.get("tags", []))
    for cf_key, cf_val in obj.get("custom_fields", {}).items():
        flat[f"CF: {cf_key}"] = safe_string(cf_val)
    return flat


# --------------------------------------------------------------------------- #
# Flatten IP Address
# --------------------------------------------------------------------------- #
def flatten_ip_address(obj: Dict[str, Any]) -> Dict[str, Any]:
    flat = {
        "ID": obj.get("id"),
        "Display": obj.get("display"),
        "Address": obj.get("address"),
        "Family": safe_string(obj.get("family")),
        "VRF": safe_string(obj.get("vrf")),
        "Tenant": safe_string(obj.get("tenant")),
        "Status": safe_string(obj.get("status")),
        "Role": safe_string(obj.get("role")),
        "Assigned To": safe_string(obj.get("assigned_object")),
        "DNS Name": obj.get("dns_name", ""),
        "Description": obj.get("description", ""),
        "Comments": obj.get("comments", ""),
        "NAT Inside": safe_string(obj.get("nat_inside")),
        "NAT Outside": safe_string(obj.get("nat_outside")),
        "Created": obj.get("created"),
        "Last Updated": obj.get("last_updated"),
    }
    flat["Tags"] = safe_string(obj.get("tags", []))
    for cf_key, cf_val in obj.get("custom_fields", {}).items():
        flat[f"CF: {cf_key}"] = safe_string(cf_val)
    return flat


# --------------------------------------------------------------------------- #
# Fetch all pages
# --------------------------------------------------------------------------- #
def fetch_all_pages(session: requests.Session, url: str) -> List[Dict[str, Any]]:
    results = []
    next_url = url
    while next_url:
        resp = session.get(next_url, verify=False)
        resp.raise_for_status()
        data = resp.json()
        results.extend(data["results"])
        next_url = data.get("next")
    return results


# --------------------------------------------------------------------------- #
# Write Excel
# --------------------------------------------------------------------------- #
def write_to_excel(records: List[Dict[str, Any]], output_path: Path, sheet_name: str) -> None:
    if not records:
        print(f"No data for {sheet_name}, skipping.")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    headers = list(records[0].keys())
    ws.append(headers)

    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.font = header_font

    for rec in records:
        row = [rec.get(h, "") for h in headers]
        ws.append(row)

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except Exception:
                pass
        adjusted_width = min(max_length + 2, 60)
        ws.column_dimensions[column].width = adjusted_width

    wb.save(output_path)
    print(f"Saved: {output_path} ({len(records)} rows)")


# --------------------------------------------------------------------------- #
# Send Email with Multiple Attachments
# --------------------------------------------------------------------------- #
def send_email(attachments: List[Path]) -> None:
    if not attachments:
        print("No files to send.")
        return

    msg = EmailMessage()
    msg["From"] = FROM_EMAIL
    msg["To"] = EMAIL_TO
    msg["Subject"] = f"NetBox Export: IP Ranges & Addresses ({len(attachments)} files)"

    body = f"""
Hello ,

Attached are the latest NetBox exports:

"""
    for f in attachments:
        body += f"- {f.name} ({f.stat().st_size // 1024} KB)\n"
    body += f"\nGenerated on: {__import__('datetime').datetime.now().strftime('%Y-%m-%d %H:%M')}\n\nRegards,\nNetBox Export Bot"
    msg.set_content(body)

    for attachment_path in attachments:
        with open(attachment_path, "rb") as f:
            file_data = f.read()
        msg.add_attachment(
            file_data,
            maintype="application",
            subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename=attachment_path.name
        )

    print(f"Sending {len(attachments)} file(s) to {EMAIL_TO} via {SMTP_SERVER}:{SMTP_PORT}...")
    try:
        with smtplib.SMTP(SMTP_SERVER, SMTP_PORT, timeout=30) as server:
            server.send_message(msg)
        print("Email sent successfully!")
        
    except Exception as e:
        print(f"Failed to send email: {e}", file=sys.stderr)
        sys.exit(1)


# --------------------------------------------------------------------------- #
# Main
# --------------------------------------------------------------------------- #
def main() -> None:
    parser = argparse.ArgumentParser(
        description="Export NetBox IP Ranges + IP Addresses → Excel → Email"
    )
    parser.add_argument("-t", "--token", required=True, help="NetBox API token")
    args = parser.parse_args()

    session = requests.Session()
    session.headers.update({
        "Authorization": f"Token {args.token}",
        "Accept": "application/json"
    })

    # === Export IP Ranges ===
    print("Fetching IP Ranges...")
    try:
        ip_ranges = fetch_all_pages(session, URL_IP_RANGES)
        flat_ranges = [flatten_ip_range(r) for r in ip_ranges]
        write_to_excel(flat_ranges, OUTPUT_IP_RANGES, "IP Ranges")
    except Exception as e:
        print(f"IP Ranges failed: {e}", file=sys.stderr)
        ip_ranges_file = None
    else:
        ip_ranges_file = OUTPUT_IP_RANGES

    # === Export IP Addresses ===
    print("Fetching IP Addresses...")
    try:
        ip_addrs = fetch_all_pages(session, URL_IP_ADDRESSES)
        flat_addrs = [flatten_ip_address(a) for a in ip_addrs]
        write_to_excel(flat_addrs, OUTPUT_IP_ADDRESSES, "IP Addresses")
    except Exception as e:
        print(f"IP Addresses failed: {e}", file=sys.stderr)
        ip_addrs_file = None
    else:
        ip_addrs_file = OUTPUT_IP_ADDRESSES

    # === Send Email ===
    attachments = [f for f in [ip_ranges_file, ip_addrs_file] if f and f.exists()]
    if attachments:
        send_email(attachments)
    else:
        print("No files generated. Email not sent.")
        sys.exit(1)


if __name__ == "__main__":
    main()
